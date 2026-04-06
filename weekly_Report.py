from cProfile import label
from logging import root
import os
import re
#import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from datetime import datetime, timedelta
from openpyxl.styles import Alignment, Font, PatternFill
from pathlib import Path
import sys
import tkinter as tk
from tkinter import messagebox
import winsound

from PIL import Image, ImageTk

# === CONFIG ===
if getattr(sys, 'frozen', False):
    # Running inside PyInstaller bundle
    PROJECT_ROOT = Path(sys._MEIPASS)
else:
    # Running from source
    PROJECT_ROOT = Path(__file__).resolve().parent

HOME = Path.home()
TEMPLATE_PATH = PROJECT_ROOT / "Template.xlsx"
OUTPUT_PATH = HOME / "Documents" / "Reports" / "Monthly_Report"
img_path = PROJECT_ROOT / "success.png"


SHEET_NAME = "Sheet1"
START_ROW = 18
START_COL = 2   # B = 2

BASE_FOLDER = HOME / "Documents" / "Reports" / "Monthly_Report"
print(BASE_FOLDER)

# === Helper functions ===
def get_week_start(date):
    days_since_sunday = (date.isoweekday() % 7)
    return date - timedelta(days=days_since_sunday)

def get_week_dates(week_start):
    # Returns the whole week (Sunday to Saturday, 0 to 6 days from Sunday)
    return [week_start + timedelta(days=i) for i in range(0, 7)]

def parse_daily_file(file_path):
    tasks = []
    with open(file_path, encoding='utf-8') as f:
        content = f.read()

    # Original regex for standard tasks
    matches = re.findall(r"- Task:\s*(.*?)\s*Status:\s*(\w+)", content, re.DOTALL)
    for task_text, status in matches:
        task_text = task_text.strip()
        # Separate Repo and Commit if this is a git log format
        if "\n  Commit:" in task_text or "\nCommit:" in task_text:
            parts = re.split(r"\n\s*Commit:\s*", task_text, maxsplit=1)
            tasks.append({"Date": None, "Task": parts[0].strip(), "Commit": parts[1].strip(), "Status": status.strip()})
        else:
            tasks.append({"Date": None, "Task": task_text, "Commit": None, "Status": status.strip()})
    
    # New regex for Git commits format (Only if at start of line to avoid duplicates)
    # This matches the legacy format (Repo : ..., Commit : ...)
    # (?m)^ ensures it only matches if it's the start of a line
    git_matches = re.findall(r"(?m)^Repo\s*:\s*(.*?)\s*Commit\s*:\s*(.*)", content)
    for repo, msg in git_matches:
        tasks.append({"Date": None, "Task": f"Repo: {repo.strip()}", "Commit": msg.strip(), "Status": "Completed"})
        
    return tasks

def generate_monthly_report(target_date=None):
    if target_date:
        today = datetime.strptime(target_date, "%Y-%m-%d").date()
    else:
        today = datetime.today().date()

    month_str = f"{today.year}_{today.month:02d}"

    # Prepare new weekly data
    current_week_start = get_week_start(today)
    week_dates = get_week_dates(current_week_start)

    week_number = get_week_number_in_month(current_week_start)
    weekly_tasks = []

    for day in week_dates:
        folder_name = f"{day.year}_{day.month:02d}"
        folder_path = os.path.join(BASE_FOLDER, folder_name)
        file_name = f"{day}.txt"
        file_path = os.path.join(folder_path, file_name)

        if os.path.exists(file_path):
            tasks = parse_daily_file(file_path)
            for t in tasks:
                t["Date"] = day.isoformat()
            weekly_tasks.extend(tasks)

    if not weekly_tasks:
        print("No tasks found for this week.")
        return

    # === Load your Excel template ===
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb[SHEET_NAME]

    # Find the last used row under B18
    last_row = START_ROW - 1
    for row in ws.iter_rows(min_row=START_ROW, max_row=ws.max_row):
        cell_val = row[START_COL - 1].value
        if cell_val:
            last_row = row[0].row

    next_row = last_row + 1

    # === Insert the Week header row ===


    # === Write the new data ===
    if weekly_tasks:
        # merge + write Week header once
        merge_start = START_COL
        merge_end = START_COL + 3 # Merge B to E

        ws.merge_cells(
            start_row=next_row,
            start_column=merge_start,
            end_row=next_row,
            end_column=merge_end
        )

        week_cell = ws.cell(row=next_row, column=merge_start)
        week_cell.value = f"Week {week_number}"
        week_cell.font = Font(bold=True, color="FFFFFF",size=16)
        week_cell.alignment = Alignment(horizontal="center", vertical="center")
        week_cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

        next_row += 1

        run_date = datetime.now().date().isoformat()  # e.g. '2025-07-09'
        ws.cell(row=7, column=7, value=f"{run_date}")

        for task in weekly_tasks:
            ws.cell(row=next_row, column=START_COL, value=task["Date"])
            ws.cell(row=next_row, column=START_COL + 1, value=task["Task"])
            ws.cell(row=next_row, column=START_COL + 2, value=task["Status"])
            if task.get("Commit"):
                ws.cell(row=next_row, column=START_COL + 3, value=task["Commit"]) # Column E
            next_row += 1
    else:
        print(f"No tasks found for Week {week_number}. Skipping.")

    # Use the last day of the work week (Thursday) for filename
    end_of_week = week_dates[-1].strftime("%Y-%m-%d")
    filename = f"{end_of_week}_Report.xlsx"
    filepath = os.path.join(OUTPUT_PATH, filename)

    wb.save(filepath)

    root = tk.Tk()
    root.title("Report Completed 🎉")
    root.geometry("400x350")
    root.configure(bg="#f0f0f0")
    #root.withdraw()
    winsound.MessageBeep(winsound.MB_ICONASTERISK)
    img = Image.open(img_path)
    #img = img.resize((150, 150))  # Optional resize
    photo = ImageTk.PhotoImage(img)
    #messagebox.showinfo("Weekly Report", "✅ Report generated successfully!")
    label = tk.Label(root, text="Weekly Report Generated Successfully!", font=("Segoe UI", 12), bg="#f0f0f0")
    label.pack(pady=10)
    tk.Label(root, image=photo, bg="#f0f0f0").pack(pady=10)
    tk.Button(root, text="Close", command=root.destroy, width=10).pack(pady=10)
    root.mainloop()

    print(f"Report saved as: {filepath}")

def get_week_number_in_month(date):
    """
    Calculate week number in month for a given date.
    E.g. returns 1 for first week, 2 for second, etc.
    """
    first_day = date.replace(day=1)
    first_sunday = get_week_start(first_day)
    week_number = ((date - first_sunday).days // 7) + 1
    return week_number

import argparse

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generate Weekly Report")
    parser.add_argument("--date", type=str, help="Target date in YYYY-MM-DD format")
    args, unknown = parser.parse_known_args()
    
    target_date = args.date
    # Allow the user to also pass just the date without --date for backward compatibility
    if target_date is None and unknown and not unknown[0].startswith('--'):
        target_date = unknown[0]
        
    if target_date:
        generate_monthly_report(target_date)
    else:
        generate_monthly_report()
