import os
import re
import pandas as pd
from datetime import datetime, timedelta
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl import load_workbook

# === CONFIG ===
BASE_FOLDER = r"D:\Wassel\Reports\Monthly_Report"

def get_week_start(date):
    days_since_sunday = (date.isoweekday() % 7)
    return date - timedelta(days=days_since_sunday)

def get_week_dates(week_start):
    return [week_start + timedelta(days=i) for i in range(5)]  # Sun to Thu

def parse_daily_file(file_path):
    tasks = []
    with open(file_path, encoding='utf-8') as f:
        content = f.read()

    matches = re.findall(r"- Task:\s*(.*?)\s*Status:\s*(\w+)", content, re.DOTALL)
    for task, status in matches:
        tasks.append({"Date": None, "Task": task.strip(), "Status": status.strip()})
    return tasks

def generate_monthly_report():
    today = datetime.today().date()
    month_str = f"{today.year}_{today.month:02d}"
    excel_file = os.path.join(BASE_FOLDER, f"Monthly_Report_{month_str}.xlsx")

    # We'll gather tasks grouped by week number
    all_data = []

    # For simplicity, we get all weeks from 1 to max weeks in month
    # Or just for the current week (modify as needed)
    # Here, as example, just current week:
    current_week_start = get_week_start(today)
    week_dates = get_week_dates(current_week_start)

    # Assign a fake week number, e.g. 1 for current week (extend for multiple weeks as needed)
    week_number = 1

    weekly_tasks = []
    for day in week_dates:
        folder_name = f"{day.year}_{day.month:02d}"
        folder_path = os.path.join(BASE_FOLDER, folder_name)
        file_name = f"{day}.txt"
        file_path = os.path.join(folder_path, file_name)

        if os.path.exists(file_path):
            tasks = parse_daily_file(file_path)
            # Add date info per task
            for t in tasks:
                t["Date"] = day.isoformat()
            weekly_tasks.extend(tasks)

    if not weekly_tasks:
        print("No tasks found for this week.")
        return

    # Insert a "Week X" row at the start
    all_data.append({"Date": "", "Task": f"Week {week_number}", "Status": ""})

    # Append all tasks of this week
    all_data.extend(weekly_tasks)

    # Convert to DataFrame
    df = pd.DataFrame(all_data)

    # If Excel exists, load old data and append, but here for simplicity just overwrite
    df.to_excel(excel_file, index=False)

    # Now open Excel with openpyxl to bold Week rows
    wb = load_workbook(excel_file)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        cell_task = row[1]  # Task column is second column (B)
        if cell_task.value and str(cell_task.value).startswith("Week "):
            for cell in row:
                cell.font = Font(bold=True)

    wb.save(excel_file)

    print(f"Report saved with week headers: {excel_file}")

if __name__ == "__main__":
    generate_monthly_report()
